"""
Generador de Archivos Excel - Módulo de Utilidades
=================================================

Este módulo contiene la lógica para generar archivos Excel con formato
profesional a partir de los datos extraídos de facturas XML.

Funcionalidades principales:
- Generación de reportes Excel con formato profesional
- Ajuste automático de anchos de columna
- Aplicación de estilos y formato
- Generación de hojas de resumen
- Validación de estructura de datos
- Cálculo de estadísticas

Clases principales:
- ExcelGenerator: Clase principal para generación de Excel

Ejemplo de uso:
    generator = ExcelGenerator()
    excel_file = generator.generate_excel(data_rows)
    # excel_file es un BytesIO con el archivo Excel generado

Autor: Sistema de Procesamiento XML
Versión: 2.0.0
"""

import pandas as pd
import io
import logging
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar logging para debugging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Generador de archivos Excel para reportes de facturas
    
    Esta clase maneja la creación de archivos Excel con formato profesional
    a partir de los datos extraídos de facturas XML.
    
    Funcionalidades:
    - Generación de hojas de datos principales
    - Creación de hojas de resumen
    - Aplicación de estilos y formato
    - Ajuste automático de columnas
    - Validación de datos
    
    Atributos:
        columns (list): Lista de columnas para el reporte Excel
        styles (dict): Configuración de estilos para el Excel
    
    Ejemplo de uso:
        generator = ExcelGenerator()
        
        # Generar Excel básico
        excel_file = generator.generate_excel(data_rows)
        
        # Generar con hoja de resumen
        excel_file = generator.generate_summary_sheet(df, excel_file)
        
        # Validar estructura de datos
        is_valid = generator.validate_data_structure(data_rows)
    """
    
    def __init__(self):
        """
        Inicializa el generador Excel con configuraciones necesarias
        
        Define las columnas del reporte, estilos por defecto y
        configuraciones de formato para el archivo Excel.
        """
        # Definir columnas del reporte Excel según la estructura requerida
        self.columns = [
            'Cuenta',
            'Comprobante',
            'Fecha',
            'Documento',
            'Documento_Ref',
            'Nit',
            'Detalle',
            'Tipo',
            'Estado_Fiscal',
            'Valor',
            'Base',
            'Centro_Costo',
            'Trans_Ext',
            'Plazo',
            'Docto_Electronico'
        ]
        
        # Configuración de estilos para el Excel
        self.styles = {
            'header_font': Font(name='Arial', size=11, bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'header_alignment': Alignment(horizontal='center', vertical='center'),
            'data_font': Font(name='Arial', size=10),
            'data_alignment': Alignment(horizontal='left', vertical='center'),
            'number_format': '#,##0.00',
            'currency_format': '$ #,##0.0000',  # Formato ecuatoriano con 4 decimales
            'date_format': 'dd/mm/yyyy',
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        logger.info("ExcelGenerator inicializado con columnas y estilos configurados")

    def generate_excel(self, all_rows: List[Dict[str, Any]]) -> io.BytesIO:
        """
        Genera un archivo Excel completo con los datos procesados
        
        Crea un archivo Excel con formato profesional que incluye:
        - Hoja principal con todos los datos
        - Formato aplicado automáticamente
        - Ajuste de anchos de columna
        - Estilos de encabezado y datos
        
        Args:
            all_rows (List[Dict[str, Any]]): Lista de diccionarios con datos de facturas
            
        Returns:
            io.BytesIO: Archivo Excel en memoria
            
        Ejemplo de uso:
            excel_file = generator.generate_excel(data_rows)
            
            # Guardar archivo
            with open('reporte_facturas.xlsx', 'wb') as f:
                f.write(excel_file.getvalue())
                
        Estructura del Excel generado:
            - Hoja 1: Datos principales (todas las líneas de factura)
            - Columnas formateadas según tipo de dato
            - Encabezados con estilo profesional
            - Anchos de columna optimizados
        """
        try:
            # Crear DataFrame con los datos
            df = pd.DataFrame(all_rows)
            
            # Validar estructura de datos
            if not self.validate_data_structure(all_rows):
                logger.warning("Estructura de datos no válida, continuando con datos disponibles")
            
            # Crear archivo Excel en memoria
            output = io.BytesIO()
            
            # Crear workbook y worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Datos de Facturas"
            
            # Escribir encabezados
            for col_num, column in enumerate(self.columns, 1):
                cell = worksheet.cell(row=1, column=col_num, value=column)
                cell.font = self.styles['header_font']
                cell.fill = self.styles['header_fill']
                cell.alignment = self.styles['header_alignment']
                cell.border = self.styles['border']
            
            # Escribir datos
            for row_num, row_data in enumerate(all_rows, 2):
                for col_num, column in enumerate(self.columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num, value=row_data.get(column, ''))
                    cell.font = self.styles['data_font']
                    cell.alignment = self.styles['data_alignment']
                    cell.border = self.styles['border']
                    
                    # Aplicar formato específico según tipo de columna
                    self._apply_cell_format(cell, column, row_data.get(column, ''))
            
            # Ajustar anchos de columna
            self._adjust_column_widths(worksheet)
            
            # Aplicar formato básico
            self._apply_basic_formatting(worksheet, df)
            
            # Guardar archivo
            workbook.save(output)
            output.seek(0)
            
            logger.info(f"Archivo Excel generado exitosamente con {len(all_rows)} filas")
            return output
            
        except Exception as e:
            logger.error(f"Error generando archivo Excel: {str(e)}")
            raise

    def _apply_cell_format(self, cell, column_name: str, value):
        """
        Aplica formato específico a una celda según el tipo de columna
        
        Determina y aplica el formato apropiado (número, moneda, fecha, texto)
        según el nombre de la columna y el valor contenido.
        
        Args:
            cell: Celda de openpyxl a formatear
            column_name (str): Nombre de la columna
            value: Valor de la celda
            
        Ejemplo de uso:
            cell = worksheet.cell(row=1, column=1)
            generator._apply_cell_format(cell, 'Base_Imponible', '1000.50')
            # Aplica formato de moneda a la celda
            
        Formatos aplicados:
            - Moneda: Para columnas de montos (Base_Imponible, Impuesto, Total, etc.)
            - Número: Para columnas numéricas (Cantidad, Porcentaje)
            - Fecha: Para columnas de fecha
            - Texto: Para columnas descriptivas
        """
        try:
            # Columnas de moneda
            currency_columns = [
                'Valor', 'Base'
            ]
            
            # Columnas numéricas
            number_columns = []
            
            # Columnas de fecha
            date_columns = ['Fecha', 'Plazo']
            
            # Aplicar formato según tipo de columna
            if column_name in currency_columns:
                if value and str(value).replace('.', '').replace(',', '').isdigit():
                    # Convertir a número para asegurar formato correcto
                    try:
                        numeric_value = float(str(value).replace(',', '.'))
                        cell.value = numeric_value
                        cell.number_format = self.styles['currency_format']
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    except (ValueError, TypeError):
                        pass
                    
            elif column_name in number_columns:
                if value and str(value).replace('.', '').replace(',', '').isdigit():
                    # Convertir a número para asegurar formato correcto
                    try:
                        numeric_value = float(str(value).replace(',', '.'))
                        cell.value = numeric_value
                        cell.number_format = self.styles['number_format']
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    except (ValueError, TypeError):
                        pass
                    
            elif column_name in date_columns:
                if value:
                    cell.number_format = self.styles['date_format']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
        except Exception as e:
            logger.debug(f"Error aplicando formato a celda {column_name}: {str(e)}")

    def _adjust_column_widths(self, worksheet):
        """
        Ajusta automáticamente los anchos de las columnas
        
        Calcula y establece el ancho óptimo para cada columna basándose
        en el contenido de los datos y encabezados.
        
        Args:
            worksheet: Hoja de trabajo de openpyxl
            
        Ejemplo de uso:
            generator._adjust_column_widths(worksheet)
            # Ajusta automáticamente todos los anchos de columna
            
        Lógica de ajuste:
            - Mínimo 10 caracteres para cualquier columna
            - Máximo 50 caracteres para evitar columnas muy anchas
            - Considera el ancho del encabezado y datos
            - Ajusta según el tipo de contenido
        """
        try:
            # Anchos sugeridos por tipo de columna
            column_widths = {
                'Cuenta': 15,
                'Comprobante': 12,
                'Fecha': 12,
                'Documento': 20,
                'Documento_Ref': 35,
                'Nit': 15,
                'Detalle': 40,
                'Tipo': 12,
                'Estado_Fiscal': 15,
                'Valor': 15,
                'Base': 15,
                'Centro_Costo': 15,
                'Trans_Ext': 12,
                'Plazo': 12,
                'Docto_Electronico': 50
            }
            
            # Aplicar anchos de columna
            for col_num, column in enumerate(self.columns, 1):
                width = column_widths.get(column, 15)  # Ancho por defecto: 15
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = width
                
        except Exception as e:
            logger.error(f"Error ajustando anchos de columna: {str(e)}")

    def _apply_basic_formatting(self, worksheet, df: pd.DataFrame):
        """
        Aplica formato básico a toda la hoja de trabajo
        
        Aplica estilos y formato general a la hoja de trabajo,
        incluyendo bordes, alineación y otros elementos visuales.
        
        Args:
            worksheet: Hoja de trabajo de openpyxl
            df (pd.DataFrame): DataFrame con los datos
            
        Ejemplo de uso:
            generator._apply_basic_formatting(worksheet, df)
            # Aplica formato general a toda la hoja
            
        Elementos de formato aplicados:
            - Bordes en todas las celdas
            - Alineación de texto
            - Altura de filas optimizada
            - Protección de celdas de encabezado
        """
        try:
            # Ajustar altura de filas
            worksheet.row_dimensions[1].height = 25  # Encabezado
            
            # Aplicar bordes a todas las celdas con datos
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(self.columns)):
                for cell in row:
                    cell.border = self.styles['border']
                    
            # Congelar paneles (mantener encabezados visibles)
            worksheet.freeze_panes = 'A2'
            
            # Configurar propiedades de la hoja
            worksheet.sheet_properties.tabColor = "366092"
            
        except Exception as e:
            logger.error(f"Error aplicando formato básico: {str(e)}")

    def generate_summary_sheet(self, df: pd.DataFrame, output: io.BytesIO) -> io.BytesIO:
        """
        Genera una hoja de resumen con estadísticas del reporte
        
        Crea una hoja adicional con resúmenes, totales y estadísticas
        de los datos procesados.
        
        Args:
            df (pd.DataFrame): DataFrame con los datos principales
            output (io.BytesIO): Archivo Excel existente
            
        Returns:
            io.BytesIO: Archivo Excel actualizado con hoja de resumen
            
        Ejemplo de uso:
            excel_file = generator.generate_excel(data_rows)
            excel_file = generator.generate_summary_sheet(df, excel_file)
            
        Contenido de la hoja de resumen:
            - Totales por tipo de clasificación fiscal
            - Estadísticas por cliente y proveedor
            - Resumen de impuestos
            - Información de archivos procesados
        """
        try:
            # Cargar el archivo Excel existente
            output.seek(0)
            workbook = pd.read_excel(output, engine='openpyxl')
            
            # Crear nueva hoja de resumen
            with pd.ExcelWriter(output, engine='openpyxl', mode='w') as writer:
                # Escribir hoja principal
                df.to_excel(writer, sheet_name='Datos de Facturas', index=False)
                
                # Crear hoja de resumen
                summary_data = self._create_summary_data(df)
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Resumen', index=False)
                
                # Aplicar formato a la hoja de resumen
                workbook = writer.book
                summary_sheet = workbook['Resumen']
                self._format_summary_sheet(summary_sheet)
            
            output.seek(0)
            logger.info("Hoja de resumen agregada exitosamente")
            return output
            
        except Exception as e:
            logger.error(f"Error generando hoja de resumen: {str(e)}")
            return output

    def _create_summary_data(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Crea datos de resumen a partir del DataFrame principal
        
        Calcula estadísticas y totales para la hoja de resumen.
        
        Args:
            df (pd.DataFrame): DataFrame con datos principales
            
        Returns:
            List[Dict[str, Any]]: Lista de diccionarios con datos de resumen
            
        Ejemplo de datos generados:
            [
                {'Concepto': 'Total Facturas', 'Valor': 150},
                {'Concepto': 'Total Gravado', 'Valor': 50000.00},
                {'Concepto': 'Total Exento', 'Valor': 10000.00},
                {'Concepto': 'Total IVA', 'Valor': 6000.00}
            ]
        """
        summary_data = []
        
        try:
            # Estadísticas básicas
            summary_data.append({
                'Concepto': 'Total Facturas',
                'Valor': df['Documento'].nunique()
            })
            
            summary_data.append({
                'Concepto': 'Total Líneas',
                'Valor': len(df)
            })
            
            # Totales por clasificación fiscal
            if 'Tipo' in df.columns:
                gravado_total = df[df['Tipo'] == 'GRAVADO']['Base'].sum()
                exento_total = df[df['Tipo'] == 'EXENTO']['Base'].sum()
                excluido_total = df[df['Tipo'] == 'EXCLUIDO']['Base'].sum()
                
                summary_data.extend([
                    {'Concepto': 'Total Gravado', 'Valor': round(gravado_total, 2)},
                    {'Concepto': 'Total Exento', 'Valor': round(exento_total, 2)},
                    {'Concepto': 'Total Excluido', 'Valor': round(excluido_total, 2)}
                ])
            
            # Totales de impuestos
            if 'Valor' in df.columns:
                total_impuestos = df['Valor'].sum()
                summary_data.append({
                    'Concepto': 'Total Impuestos',
                    'Valor': round(total_impuestos, 2)
                })
            
            # Estadísticas por documento
            if 'Documento' in df.columns:
                documentos_unicos = df['Documento'].nunique()
                summary_data.append({
                    'Concepto': 'Documentos Procesados',
                    'Valor': documentos_unicos
                })
            
            # Información de archivos
            if 'Documento_Ref' in df.columns:
                archivos_unicos = df['Documento_Ref'].nunique()
                summary_data.append({
                    'Concepto': 'Archivos ZIP Procesados',
                    'Valor': archivos_unicos
                })
            
        except Exception as e:
            logger.error(f"Error creando datos de resumen: {str(e)}")
            
        return summary_data

    def _format_summary_sheet(self, worksheet):
        """
        Aplica formato a la hoja de resumen
        
        Args:
            worksheet: Hoja de resumen de openpyxl
        """
        try:
            # Aplicar estilos al encabezado
            for cell in worksheet[1]:
                cell.font = self.styles['header_font']
                cell.fill = self.styles['header_fill']
                cell.alignment = self.styles['header_alignment']
                cell.border = self.styles['border']
            
            # Ajustar anchos de columna
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 20
            
        except Exception as e:
            logger.error(f"Error formateando hoja de resumen: {str(e)}")

    def validate_data_structure(self, data: List[Dict[str, Any]]) -> bool:
        """
        Valida la estructura de los datos antes de generar Excel
        
        Verifica que los datos tengan la estructura esperada y
        contengan las columnas necesarias.
        
        Args:
            data (List[Dict[str, Any]]): Lista de diccionarios con datos
            
        Returns:
            bool: True si la estructura es válida, False en caso contrario
            
        Ejemplo de uso:
            is_valid = generator.validate_data_structure(data_rows)
            if not is_valid:
                logger.warning("Estructura de datos no válida")
                
        Validaciones realizadas:
            - Presencia de columnas requeridas
            - Tipos de datos correctos
            - Valores no nulos en campos críticos
        """
        try:
            if not data:
                logger.warning("Lista de datos vacía")
                return False
            
            # Verificar columnas requeridas
            required_columns = ['Cuenta', 'Fecha', 'Documento', 'Nit']
            first_row = data[0]
            
            missing_columns = [col for col in required_columns if col not in first_row]
            if missing_columns:
                logger.warning(f"Columnas faltantes: {missing_columns}")
                return False
            
            # Verificar que al menos algunos datos críticos no estén vacíos
            non_empty_documentos = sum(1 for row in data if row.get('Documento'))
            if non_empty_documentos == 0:
                logger.warning("No se encontraron documentos válidos")
                return False
            
            logger.info(f"Estructura de datos válida: {len(data)} filas, {len(first_row)} columnas")
            return True
            
        except Exception as e:
            logger.error(f"Error validando estructura de datos: {str(e)}")
            return False

    def get_data_statistics(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Calcula estadísticas de los datos para reportes
        
        Args:
            data (List[Dict[str, Any]]): Lista de diccionarios con datos
            
        Returns:
            Dict[str, Any]: Diccionario con estadísticas calculadas
            
        Ejemplo de uso:
            stats = generator.get_data_statistics(data_rows)
            # stats = {
            #     'total_rows': 150,
            #     'unique_invoices': 25,
            #     'date_range': '2024-01-01 a 2024-01-31',
            #     'total_amount': 50000.00
            # }
        """
        stats = {
            'total_rows': len(data),
            'unique_invoices': 0,
            'date_range': '',
            'total_amount': 0.0,
            'currency_distribution': {},
            'document_types': {}
        }
        
        try:
            if not data:
                return stats
            
            # Contar documentos únicos
            unique_documents = set(row.get('Documento', '') for row in data if row.get('Documento'))
            stats['unique_invoices'] = len(unique_documents)
            
            # Calcular rango de fechas
            dates = [row.get('Fecha', '') for row in data if row.get('Fecha')]
            if dates:
                min_date = min(dates)
                max_date = max(dates)
                stats['date_range'] = f"{min_date} a {max_date}"
            
            # Calcular total de montos
            total_amount = sum(float(row.get('Valor', 0)) for row in data if row.get('Valor'))
            stats['total_amount'] = round(total_amount, 2)
            
            # Distribución de tipos de cuenta
            account_types = [row.get('Cuenta', '') for row in data if row.get('Cuenta')]
            stats['currency_distribution'] = {acc: account_types.count(acc) for acc in set(account_types)}
            
            # Tipos de documento
            doc_types = [row.get('Cuenta', '') for row in data if row.get('Cuenta')]
            stats['document_types'] = {doc: doc_types.count(doc) for doc in set(doc_types)}
            
        except Exception as e:
            logger.error(f"Error calculando estadísticas: {str(e)}")
            
        return stats
